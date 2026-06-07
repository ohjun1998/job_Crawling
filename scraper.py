import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import urllib.parse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 브라우저 기본 헤더 설정 (PC 버전으로 통합)
PC_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
    'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8'
}

# 무관한 직무를 걸러내기 위한 블랙리스트 키워드
EXCLUDE_KEYWORDS = [
    "요양", "주간보호", "주야간보호", "보호소", "보호센터", "사회복지", "복지사", 
    "요양보호사", "간호", "물리치료", "작업치료", "치료사", "조리사", "영양사", 
    "노인", "어르신", "케어", "재활", "교육환경보호원", "아동보호", "교사", "운전원", "사무원"
]

def should_exclude(title, corp):
    """제외 키워드 포함 여부 검사"""
    for kw in EXCLUDE_KEYWORDS:
        if kw in title or kw in corp:
            return True
    return False

def scrape_saramin(keyword):
    """1. 사람인 채용공고 크롤링"""
    jobs = []
    url = f"https://www.saramin.co.kr/zf_user/search/recruit?searchword={keyword}"
    try:
        res = requests.get(url, headers=PC_HEADERS, timeout=8)
        if res.status_code == 200:
            soup = BeautifulSoup(res.text, 'html.parser')
            items = soup.select('div.item_recruit')
            for item in items:
                try:
                    title_elem = item.select_one('h2.job_tit a')
                    corp_elem = item.select_one('div.area_corp strong a') or item.select_one('div.area_corp a')
                    if title_elem and corp_elem:
                        title, corp = title_elem.get_text(strip=True), corp_elem.get_text(strip=True)
                        if should_exclude(title, corp): continue
                        jobs.append({"사이트": "사람인", "회사명": corp, "제목": title, "링크": "https://www.saramin.co.kr" + title_elem['href']})
                except Exception: continue
            print(f"  └ [사람인] '{keyword}' -> 필터링 후 {len(jobs)}건 수집 완료")
    except Exception: print(f"  └ [사람인] 연결 실패 ({keyword})")
    return jobs

def scrape_jobkorea(keyword):
    """2. 잡코리아 최신 개편 레이아웃 대응 크롤링"""
    jobs = []
    url = f"https://www.jobkorea.co.kr/Search/?stext={keyword}"
    try:
        res = requests.get(url, headers=PC_HEADERS, timeout=8)
        print(f"  └ [잡코리아] '{keyword}' 접속 -> HTTP 상태코드: {res.status_code}")
        
        if res.status_code == 200:
            soup = BeautifulSoup(res.text, 'html.parser')
            
            # 보내주신 HTML의 핵심 앵커인 data-sentry-component="Title" 속성 추적
            title_elements = soup.select('a[data-sentry-component="Title"]')
            print(f"  └ [잡코리아] 발견된 신규 레이아웃 공고 수: {len(title_elements)}개")
            
            # 시나리오 A: 신규 레이아웃 요소가 발견된 경우 (정상 파싱)
            if len(title_elements) > 0:
                for title_elem in title_elements:
                    try:
                        title = title_elem.get_text(strip=True)
                        link = title_elem['href']
                        if not link.startswith('http'):
                            link = "https://www.jobkorea.co.kr" + link
                        
                        # 상위 카드 컴포넌트(w-full)로 이동하여 내부의 회사명 클래스(text-typo-b2-16) 추출
                        parent_card = title_elem.find_parent('div', class_='w-full') or title_elem.find_parent('div')
                        corp_elem = parent_card.select_one('span.text-typo-b2-16') if parent_card else None
                        corp = corp_elem.get_text(strip=True) if corp_elem else "회사명 확인불가"
                        
                        if should_exclude(title, corp): continue
                        jobs.append({"사이트": "잡코리아", "회사명": corp, "제목": title, "링크": link})
                    except Exception: continue
            
            # 시나리오 B: 만약 신규 레이아웃이 안 잡힐 경우를 대비한 구형 백업 선택자 가동
            else:
                backup_items = soup.select('li.list-post') or soup.select('tr.dev_item')
                for item in backup_items:
                    try:
                        title_elem = item.select_one('div.post-list-info a.title') or item.select_one('.tit a')
                        corp_elem = item.select_one('div.post-list-corp a') or item.select_one('.corp a')
                        if title_elem and corp_elem:
                            title, corp = title_elem.get_text(strip=True), corp_elem.get_text(strip=True)
                            if should_exclude(title, corp): continue
                            link = title_elem['href']
                            if not link.startswith('http'): link = "https://www.jobkorea.co.kr" + link
                            jobs.append({"사이트": "잡코리아", "회사명": corp, "제목": title, "링크": link})
                    except Exception: continue
                    
            print(f"  └ [잡코리아] 최종 필터링 후 {len(jobs)}건 저장 대기")
        else:
            print(f"  └ [잡코리아] 방화벽 거부 (Status: {res.status_code}) -> 로컬 실행 권장")
    except Exception as e: 
        print(f"  └ [잡코리아] 타임아웃 또는 예외 발생: {e}")
    return jobs

def scrape_incruit(keyword):
    """3. 인크루트 채용공고 크롤링"""
    jobs = []
    url = f"https://search.incruit.com/list/search.asp?col=job&kw={keyword}"
    try:
        res = requests.get(url, headers=PC_HEADERS, timeout=8)
        if res.status_code == 200:
            soup = BeautifulSoup(res.text, 'html.parser')
            items = soup.select('div.cl_md ul.docs_list li') or soup.select('ul.win_list li')
            for item in items:
                try:
                    title_elem = item.select_one('span.txt_tit a') or item.select_one('p.tit a')
                    corp_elem = item.select_one('span.txt_corp a') or item.select_one('p.cpname a')
                    if title_elem and corp_elem:
                        title, corp = title_elem.get_text(strip=True), corp_elem.get_text(strip=True)
                        if should_exclude(title, corp): continue
                        jobs.append({"사이트": "인크루트", "회사명": corp, "제목": title, "링크": title_elem['href']})
                except Exception: continue
            print(f"  └ [인크루트] '{keyword}' -> 필터링 후 {len(jobs)}건 수집 완료")
    except Exception: print(f"  └ [인크루트] 연결 실패 ({keyword})")
    return jobs

def scrape_remember(keyword):
    """4. 리멤버 채용공고 크롤링"""
    jobs = []
    url = f"https://career.rememberapp.com/job/search?keyword={keyword}"
    try:
        res = requests.get(url, headers=PC_HEADERS, timeout=5)
        if res.status_code == 200:
            soup = BeautifulSoup(res.text, 'html.parser')
            items = soup.select('div[class*="JobCard"]') or soup.select('article')
            for item in items:
                try:
                    title_elem = item.select_one('h3') or item.select_one('[class*="Title"]')
                    corp_elem = item.select_one('span[class*="CompanyName"]') or item.select_one('[class*="Company"]')
                    link_elem = item.select_one('a')
                    if title_elem and corp_elem and link_elem:
                        title, corp = title_elem.get_text(strip=True), corp_elem.get_text(strip=True)
                        if should_exclude(title, corp): continue
                        jobs.append({"사이트": "리멤버", "회사명": corp, "제목": title, "링크": "https://career.rememberapp.com" + link_elem['href']})
                except Exception: continue
            print(f"  └ [리멤버] '{keyword}' -> 필터링 후 {len(jobs)}건 수집 완료")
    except Exception: pass
    return jobs

def scrape_google(keyword):
    """5. 구글 검색결과 크롤링"""
    jobs = []
    query = urllib.parse.quote(f"{keyword} 채용공고")
    url = f"https://www.google.com/search?q={query}"
    try:
        res = requests.get(url, headers=PC_HEADERS, timeout=5)
        if res.status_code == 200:
            soup = BeautifulSoup(res.text, 'html.parser')
            items = soup.select('div.g') or soup.select('div.tF2Cxc')
            for item in items:
                try:
                    title_elem = item.select_one('h3')
                    link_elem = item.select_one('a')
                    if title_elem and link_elem:
                        title = title_elem.get_text(strip=True)
                        if should_exclude(title, ""): continue
                        jobs.append({"사이트": "구글검색", "회사명": "검색결과참조", "제목": title, "リンク": link_elem['href']})
                except Exception: continue
            print(f"  └ [구글검색] '{keyword}' -> 필터링 후 {len(jobs)}건 수집 완료")
    except Exception: pass
    return jobs

def main():
    keywords = ["취약점", "모의해킹", "보안", "보호", "security"]
    all_jobs = []
    
    for kw in keywords:
        print(f"--- 통합 검색 진행 중: {kw} ---")
        all_jobs.extend(scrape_saramin(kw))
        all_jobs.extend(scrape_jobkorea(kw))
        all_jobs.extend(scrape_incruit(kw))
        all_jobs.extend(scrape_remember(kw))
        all_jobs.extend(scrape_google(kw))
        time.sleep(1.5)
    
    print("\n==================================================")
    if all_jobs:
        df = pd.DataFrame(all_jobs)
        df.drop_duplicates(subset=['링크' if '링크' in df.columns else '링크'], inplace=True)
        
        output_filename = "job_results.xlsx"
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="보안 채용공고")
            workbook = writer.book
            worksheet = writer.sheets["보안 채용공고"]
            
            header_fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
            header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
            data_font = Font(name="맑은 고딕", size=10, color="333333")
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            thin_side = Side(style='thin', color='D9D9D9')
            grid_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            
            worksheet.row_dimensions[1].height = 26
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
                cell.border = grid_border
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                worksheet.row_dimensions[row[0].row].height = 20
                for cell in row:
                    cell.font = data_font
                    cell.border = grid_border
                    cell.alignment = align_center if cell.column_letter == 'A' else align_left
            
            for col in worksheet.columns:
                col_letter = col[0].column_letter
                if col_letter == 'D':
                    worksheet.column_dimensions[col_letter].width = 45
                else:
                    max_len = 0
                    for cell in col:
                        val = str(cell.value or '')
                        char_length = sum(2 if ord(char) > 128 else 1 for char in val)
                        if char_length > max_len: max_len = char_length
                    worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                    
        print(f"🎉 엑셀 최적화 완료! 파일 이름: {output_filename}")
    else:
        print("😭 수집 및 필터링된 공고가 전혀 없습니다.")
    print("==================================================")

if __name__ == "__main__":
    main()
